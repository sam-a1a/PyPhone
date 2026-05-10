import pygame
import os
from datetime import datetime
from config import SCREEN_WIDTH, SCREEN_HEIGHT, COLOR_WHITE, COLOR_BLACK
from apps.shared import Database

BLUE = (0, 122, 255)
RED = (255, 59, 48)
GREEN = (52, 199, 89)
ORANGE = (255, 149, 0)
PURPLE = (175, 82, 222)
PINK = (255, 45, 85)
GREY = (142, 142, 147)
LIGHT_BG = (245, 245, 250)

FONT_LARGE = pygame.font.SysFont("Arial", 28, bold=True)
FONT_MEDIUM = pygame.font.SysFont("Arial", 18, bold=True)
FONT_BODY = pygame.font.SysFont("Arial", 16)
FONT_SMALL = pygame.font.SysFont("Arial", 14)
FONT_TINY = pygame.font.SysFont("Arial", 12)
FONT_BUTTON = pygame.font.SysFont("Arial", 17, bold=True)

NAVBAR_HEIGHT = 85

MESSAGE_DURATION = 3000

CURRENT_FILE = os.path.abspath(__file__)
SCREENS_DIR = os.path.dirname(CURRENT_FILE)
ADMIN_DIR = os.path.dirname(SCREENS_DIR)
APPS_DIR = os.path.dirname(ADMIN_DIR)
PROJECT_ROOT = os.path.dirname(APPS_DIR)

ICON_DIR = os.path.join(PROJECT_ROOT, "assets", "icons")
REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports")


class ReportsScreen:

    def __init__(self):
        self.db = Database()
        self.background_color = LIGHT_BG

        self.message = None
        self.message_type = None  # "success" or "error"
        self.message_time = 0
        self.saved_file_path = None

        self.button_rects = {}
        self.report_cards = []

        self.scroll_y = 0
        self.max_scroll = 200

        self.icon_cache = {}

        if not os.path.exists(REPORTS_DIR):
            try:
                os.makedirs(REPORTS_DIR)
                print(f"Created reports directory at: {REPORTS_DIR}")
            except OSError as e:
                print(f"Error creating directory: {e}")

    def _get_icon(self, icon_name, size=(20, 20)):
        if icon_name in self.icon_cache:
            return self.icon_cache[icon_name]

        icon_path = os.path.join(ICON_DIR, icon_name)

        if os.path.exists(icon_path):
            try:
                img = pygame.image.load(icon_path).convert_alpha()
                img = pygame.transform.smoothscale(img, size)
                self.icon_cache[icon_name] = img
                return img
            except pygame.error:
                pass
        return None

    def handle_event(self, event):
        if event.type == pygame.MOUSEBUTTONDOWN:
            x, y = event.pos

            for card_rect, report_type, format_type in self.report_cards:
                if card_rect.collidepoint(x, y):
                    self.generate_report(report_type, format_type)
                    return None

        elif event.type == pygame.MOUSEWHEEL:
            self.scroll_y -= event.y * 30
            self.scroll_y = max(0, min(self.max_scroll, self.scroll_y))

        return None

    def update(self):
        if self.message and pygame.time.get_ticks() - self.message_time > MESSAGE_DURATION:
            self.message = None
            self.message_type = None
            self.saved_file_path = None

    def generate_report(self, report_type, file_format):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            if file_format == "pdf":
                success, filepath = self._generate_pdf_report(report_type, timestamp)
            else:  # excel
                success, filepath = self._generate_excel_report(report_type, timestamp)

            if success:
                filename = os.path.basename(filepath)
                self.message = f"✓ Saved: {filename}"
                self.message_type = "success"
                self.saved_file_path = filepath
                print(f"SUCCESS: Report saved to {filepath}")
            else:
                if not self.message:
                    self.message = f"✗ Failed to generate {file_format.upper()}"
                    self.message_type = "error"

        except Exception as e:
            print(f"CRITICAL ERROR: {e}")
            self.message = f"✗ Error: {str(e)[:20]}..."
            self.message_type = "error"

        self.message_time = pygame.time.get_ticks()

    def _generate_pdf_report(self, report_type, timestamp):
        #Generate a PDF report using reportlab
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            self.message = "✗ Install 'reportlab' library!"
            self.message_type = "error"
            print("ERROR: reportlab not installed. Run: pip install reportlab")
            return False, None

        filename = f"{report_type}_report_{timestamp}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)

        try:
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=50, leftMargin=50,
                topMargin=50, bottomMargin=50
            )

            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24, spaceAfter=20, alignment=TA_CENTER,
                textColor=colors.HexColor('#0A84FF')
            )

            # Content
            title_text = report_type.title() + " Report"
            elements.append(Paragraph(title_text, title_style))

            date_str = datetime.now().strftime("%B %d, %Y at %H:%M")
            elements.append(Paragraph(f"Generated on {date_str}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Table Data
            data = self._get_report_data(report_type)

            if data and len(data) > 1:
                table = Table(data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0A84FF')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5FA')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E5EA')),
                ]))
                elements.append(table)
            else:
                elements.append(Paragraph("No data available for this report.", styles['Normal']))

            doc.build(elements)
            return True, filepath

        except Exception as e:
            print(f"PDF Generation Error: {e}")
            self.message = "✗ PDF Error (Check Console)"
            self.message_type = "error"
            return False, None

    def _generate_excel_report(self, report_type, timestamp):
        #Generate an Excel report using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.message = "✗ Install 'openpyxl' library!"
            self.message_type = "error"
            print("ERROR: openpyxl not installed")
            return False, None

        filename = f"{report_type}_report_{timestamp}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = report_type.title()[:30]

            data = self._get_report_data(report_type)
            if not data:
                data = [["No Data"], ["There are no records to display."]]

            for col, val in enumerate(data[0], 1):
                cell = ws.cell(row=1, column=col, value=val)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0A84FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for r_idx, row in enumerate(data[1:], 2):
                for c_idx, val in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

            wb.save(filepath)
            return True, filepath

        except Exception as e:
            print(f"Excel Generation Error: {e}")
            self.message = "✗ Excel Error (Check Console)"
            self.message_type = "error"
            return False, None

    def _get_report_data(self, report_type):
        if report_type == "doctors":
            doctors = self.db.get_all_doctors(active_only=False)
            data = [["Name", "Email", "Specialty", "Phone", "Patients"]]
            for doc in doctors:
                p_count = self.db.get_doctor_patient_count(doc.id)
                data.append([doc.name, doc.email, doc.specialty, doc.phone, str(p_count)])

        elif report_type == "patients":
            patients = self.db.get_all_patients()
            data = [["Name", "Email", "Age", "Condition", "Doctor"]]
            for p in patients:
                doc = self.db.get_doctor(p.assigned_doctor_id)
                doc_name = doc.name if doc else "None"
                data.append([p.name, p.email, str(p.age), p.disease, doc_name])

        elif report_type == "appointments":
            appts = self.db.get_all_appointments()
            data = [["Date", "Time", "Patient", "Doctor", "Status"]]
            for a in appts:
                d_str = a.appointment_date.strftime("%Y-%m-%d") if a.appointment_date else ""
                data.append([d_str, a.appointment_time, getattr(a, 'patient_name', ''),
                             getattr(a, 'doctor_name', ''), a.status])

        elif report_type == "statistics":
            stats = self.db.get_statistics()
            data = [["Metric", "Value"]]
            for k, v in stats.items():
                data.append([k.replace("_", " ").title(), str(v)])
        else:
            data = []

        return data

    def draw(self, screen):
        screen.fill(self.background_color)
        self.report_cards = []

        title = FONT_LARGE.render("Reports", True, COLOR_BLACK)
        screen.blit(title, (20, 55))

        subtitle = FONT_SMALL.render("Generate PDF and Excel reports", True, GREY)
        screen.blit(subtitle, (20, 90))

        message_height = 0
        if self.message:
            message_height = self._draw_message(screen, 120)

        cards_start_y = 130 + message_height
        cards_y = cards_start_y - self.scroll_y
        card_height = 140
        card_gap = 15

        reports = [
            {"type": "doctors", "title": "Doctors Report", "description": "List of all medical staff.", "icon": "👨‍⚕️", "color": BLUE},
            {"type": "patients", "title": "Patients Report", "description": "Patient records details.", "icon": "🏥", "color": GREEN},
            {"type": "appointments", "title": "Appointments", "description": "All scheduled visits.", "icon": "📅", "color": ORANGE},
            {"type": "statistics", "title": "Statistics", "description": "Overall hospital metrics.", "icon": "📊", "color": PURPLE}
        ]

        for i, report in enumerate(reports):
            y_pos = cards_y + (i * (card_height + card_gap))
            if 100 < y_pos < SCREEN_HEIGHT - NAVBAR_HEIGHT - 20:
                self._draw_report_card(screen, report, 15, y_pos, card_height)

        total_h = len(reports) * (card_height + card_gap)
        self.max_scroll = max(0, total_h - (SCREEN_HEIGHT - NAVBAR_HEIGHT - cards_start_y))

    def _draw_message(self, screen, y):
        if not self.message: return 0
        msg_height = 50
        msg_bg = (220, 255, 220) if self.message_type == "success" else (255, 220, 220)
        msg_border = GREEN if self.message_type == "success" else RED

        rect = pygame.Rect(15, y, SCREEN_WIDTH - 30, 50)
        pygame.draw.rect(screen, msg_bg, rect, border_radius=12)
        pygame.draw.rect(screen, msg_border, rect, 2, border_radius=12)

        txt = FONT_BODY.render(self.message, True, msg_border)
        screen.blit(txt, (30, y + (50 - txt.get_height()) // 2))
        return 60

    def _draw_report_card(self, screen, report, x, y, height):
        width = SCREEN_WIDTH - 30
        pygame.draw.rect(screen, COLOR_WHITE, (x, y, width, height), border_radius=20)
        pygame.draw.rect(screen, report["color"], (x, y, 6, height), border_top_left_radius=20, border_bottom_left_radius=20)

        try:
            icon_font = pygame.font.SysFont("Segoe UI Emoji", 28)
        except:
            icon_font = pygame.font.SysFont("Arial", 28)

        icon_surf = icon_font.render(report["icon"], True, report["color"])
        screen.blit(icon_surf, (x + 25, y + 20))

        title_surf = FONT_MEDIUM.render(report["title"], True, COLOR_BLACK)
        screen.blit(title_surf, (x + 70, y + 22))

        desc_lines = self._wrap_text(report["description"], FONT_SMALL, width - 40)
        for i, line in enumerate(desc_lines[:2]):
            line_surf = FONT_SMALL.render(line, True, GREY)
            screen.blit(line_surf, (x + 25, y + 45 + (i * 18)))

        btn_y = y + height - 45

        pdf_rect = pygame.Rect(x + 25, btn_y, 100, 32)
        pygame.draw.rect(screen, RED, pdf_rect, border_radius=8)
        self.report_cards.append((pdf_rect, report["type"], "pdf"))

        pdf_img = self._get_icon("pdf.png")
        if pdf_img:
            screen.blit(pdf_img, (pdf_rect.x + 10, pdf_rect.y + 6))
            lbl_x = pdf_rect.x + 35
        else:
            lbl_x = pdf_rect.x + 35

        pdf_lbl = FONT_SMALL.render("PDF", True, COLOR_WHITE)
        screen.blit(pdf_lbl, (lbl_x, pdf_rect.y + 8))

        excel_rect = pygame.Rect(x + 135, btn_y, 100, 32)
        pygame.draw.rect(screen, GREEN, excel_rect, border_radius=8)
        self.report_cards.append((excel_rect, report["type"], "excel"))

        excel_img = self._get_icon("excel.png")
        if excel_img:
            screen.blit(excel_img, (excel_rect.x + 10, excel_rect.y + 6))
            lbl_x = excel_rect.x + 35
        else:
            lbl_x = excel_rect.x + 35

        excel_lbl = FONT_SMALL.render("Excel", True, COLOR_WHITE)
        screen.blit(excel_lbl, (lbl_x, excel_rect.y + 8))

    def _wrap_text(self, text, font, max_width):
        words = text.split(' ')
        lines = []
        current_line = []
        for word in words:
            test_line = ' '.join(current_line + [word])
            if font.size(test_line)[0] <= max_width:
                current_line.append(word)
            else:
                if current_line: lines.append(' '.join(current_line))
                current_line = [word]
        if current_line: lines.append(' '.join(current_line))
        return lines